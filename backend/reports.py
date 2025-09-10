import io
from flask import send_file
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook

from models import Simulation, TrafficData, db

# Lane mapping 
LANE_LABELS = {
    "north_in_0": "Anzac Parade (Northbound - Lane 1)",
    "north_in_1": "Anzac Parade (Northbound - Lane 2)",
    "south_in_0": "Anzac Parade (Southbound - Lane 1)",
    "south_in_1": "Anzac Parade (Southbound - Lane 2)",
    "east_in_0": "Alison Road (Eastbound - Lane 1)",
    "east_in_1": "Alison Road (Eastbound - Lane 2)",
    "west_in_0": "Alison Road (Westbound - Lane 1)",
    "west_in_1": "Alison Road (Westbound - Lane 2)",
}


def generate_simulation_pdf(sim_id: int):
    buffer = io.BytesIO()
    sim = Simulation.query.get(sim_id)
    if not sim:
        return None

    rows = TrafficData.query.filter_by(simulation_id=sim_id).all()

    # Aggregate by lane
    lane_stats = {}
    for r in rows:
        lane_stats.setdefault(r.lane, {"vehicles": 0, "queues": [], "speeds": [], "emergencies": 0})
        lane_stats[r.lane]["vehicles"] += r.vehicle_count
        lane_stats[r.lane]["queues"].append(r.queue_length)
        lane_stats[r.lane]["speeds"].append(r.avg_speed)
        if r.emergency:
            lane_stats[r.lane]["emergencies"] += 1

    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    elements = []
    elements.append(Paragraph(f"Simulation Report - ID {sim.id}", styles["Title"]))
    elements.append(Paragraph(f"Start: {sim.start_time}", styles["Normal"]))
    elements.append(Paragraph(f"End: {sim.end_time or 'Ongoing'}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Summary Table
    data = [["Lane", "Vehicles", "Avg Queue", "Avg Speed", "Emergencies"]]
    for lane, stats in lane_stats.items():
        avg_queue = sum(stats["queues"]) / len(stats["queues"]) if stats["queues"] else 0
        avg_speed = sum(stats["speeds"]) / len(stats["speeds"]) if stats["speeds"] else 0
        data.append([
            LANE_LABELS.get(lane, lane),
            stats["vehicles"],
            round(avg_queue, 2),
            round(avg_speed, 2),
            stats["emergencies"],
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return buffer


def generate_simulation_excel(sim_id: int):
    sim = Simulation.query.get(sim_id)
    if not sim:
        return None

    rows = TrafficData.query.filter_by(simulation_id=sim_id).order_by(TrafficData.timestamp).all()

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    # Raw Data Sheet (keep lane code as-is)
    ws_raw.append(["timestamp", "lane", "vehicle_count", "queue_length", "avg_speed", "emergency"])
    for r in rows:
        ws_raw.append([r.timestamp.isoformat(), r.lane, r.vehicle_count, r.queue_length, r.avg_speed, r.emergency])

    # Aggregate Sheet (use mapped labels)
    ws_agg = wb.create_sheet("Aggregated")
    ws_agg.append(["Lane", "Total Vehicles", "Avg Queue", "Avg Speed", "Emergencies"])

    lane_stats = {}
    for r in rows:
        lane_stats.setdefault(r.lane, {"vehicles": 0, "queues": [], "speeds": [], "emergencies": 0})
        lane_stats[r.lane]["vehicles"] += r.vehicle_count
        lane_stats[r.lane]["queues"].append(r.queue_length)
        lane_stats[r.lane]["speeds"].append(r.avg_speed)
        if r.emergency:
            lane_stats[r.lane]["emergencies"] += 1

    for lane, stats in lane_stats.items():
        avg_queue = sum(stats["queues"]) / len(stats["queues"]) if stats["queues"] else 0
        avg_speed = sum(stats["speeds"]) / len(stats["speeds"]) if stats["speeds"] else 0
        ws_agg.append([
            LANE_LABELS.get(lane, lane),
            stats["vehicles"],
            round(avg_queue, 2),
            round(avg_speed, 2),
            stats["emergencies"],
        ])

    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Simulation ID", sim.id])
    ws_summary.append(["Start", sim.start_time.isoformat()])
    ws_summary.append(["End", sim.end_time.isoformat() if sim.end_time else "Ongoing"])
    ws_summary.append([])
    ws_summary.append(["Total Records", len(rows)])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_comparison_excel(sim_ids: list[int]):
    """Generate Excel comparing multiple simulations by aggregates."""
    wb = Workbook()

    # Sheet: Summary
    ws_summary = wb.active
    ws_summary.title = "Comparison Summary"
    ws_summary.append(["Simulation ID", "Start", "End", "Total Vehicles", "Avg Queue", "Avg Speed", "Emergencies"])

    for sim_id in sim_ids:
        sim = Simulation.query.get(sim_id)
        if not sim:
            continue

        rows = TrafficData.query.filter_by(simulation_id=sim_id).all()

        total_vehicles = sum(r.vehicle_count for r in rows)
        avg_queue = sum(r.queue_length for r in rows) / len(rows) if rows else 0
        avg_speed = sum(r.avg_speed for r in rows) / len(rows) if rows else 0
        emergencies = sum(1 for r in rows if r.emergency)

        ws_summary.append([
            sim.id,
            sim.start_time.isoformat(),
            sim.end_time.isoformat() if sim.end_time else "Ongoing",
            total_vehicles,
            round(avg_queue, 2),
            round(avg_speed, 2),
            emergencies,
        ])

    # Sheet: Per-Lane Stats (use mapped labels)
    ws_lane = wb.create_sheet("Per-Lane Stats")
    ws_lane.append(["Simulation ID", "Lane", "Total Vehicles", "Avg Queue", "Avg Speed", "Emergencies"])

    for sim_id in sim_ids:
        sim = Simulation.query.get(sim_id)
        if not sim:
            continue

        rows = TrafficData.query.filter_by(simulation_id=sim_id).all()
        lane_stats = {}

        for r in rows:
            lane_stats.setdefault(r.lane, {"vehicles": 0, "queues": [], "speeds": [], "emergencies": 0})
            lane_stats[r.lane]["vehicles"] += r.vehicle_count
            lane_stats[r.lane]["queues"].append(r.queue_length)
            lane_stats[r.lane]["speeds"].append(r.avg_speed)
            if r.emergency:
                lane_stats[r.lane]["emergencies"] += 1

        for lane, stats in lane_stats.items():
            avg_queue = sum(stats["queues"]) / len(stats["queues"]) if stats["queues"] else 0
            avg_speed = sum(stats["speeds"]) / len(stats["speeds"]) if stats["speeds"] else 0
            ws_lane.append([
                sim_id,
                LANE_LABELS.get(lane, lane),
                stats["vehicles"],
                round(avg_queue, 2),
                round(avg_speed, 2),
                stats["emergencies"],
            ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
